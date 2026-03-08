"""
Generate a Control Assessment Worksheet (CAW) spreadsheet for NIST audits.

Usage:
    python generate_assessment_worksheet.py --framework 800-53 --baseline moderate --output caw.xlsx

Generates a pre-populated worksheet with controls for the selected framework and baseline,
ready for the auditor to fill in assessment results.
"""

import json
import sys
import argparse
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# Status options for data validation
STATUS_OPTIONS = [
    "Implemented",
    "Partially Implemented",
    "Planned",
    "Alternative Implementation",
    "Not Implemented",
    "Not Assessed",
]

EFFECTIVENESS_OPTIONS = [
    "5 - Very Effective",
    "4 - Effective",
    "3 - Moderately Effective",
    "2 - Minimally Effective",
    "1 - Ineffective",
    "0 - Non-Existent",
]

MATURITY_OPTIONS = [
    "1 - Initial/Ad Hoc",
    "2 - Repeatable",
    "3 - Defined",
    "4 - Managed",
    "5 - Optimized",
]

# 800-53 control families with representative controls per baseline
CONTROL_FAMILIES_800_53 = {
    "AC": {"name": "Access Control", "low": ["AC-1","AC-2","AC-3","AC-7","AC-8","AC-14","AC-17","AC-18","AC-19","AC-20","AC-22"], "moderate": ["AC-1","AC-2","AC-3","AC-4","AC-5","AC-6","AC-7","AC-8","AC-11","AC-12","AC-14","AC-17","AC-18","AC-19","AC-20","AC-22"], "high": ["AC-1","AC-2","AC-3","AC-4","AC-5","AC-6","AC-7","AC-8","AC-10","AC-11","AC-12","AC-14","AC-17","AC-18","AC-19","AC-20","AC-22"]},
    "AT": {"name": "Awareness and Training", "low": ["AT-1","AT-2","AT-3","AT-4"], "moderate": ["AT-1","AT-2","AT-3","AT-4"], "high": ["AT-1","AT-2","AT-3","AT-4"]},
    "AU": {"name": "Audit and Accountability", "low": ["AU-1","AU-2","AU-3","AU-4","AU-5","AU-6","AU-8","AU-9","AU-11","AU-12"], "moderate": ["AU-1","AU-2","AU-3","AU-4","AU-5","AU-6","AU-7","AU-8","AU-9","AU-11","AU-12"], "high": ["AU-1","AU-2","AU-3","AU-4","AU-5","AU-6","AU-7","AU-8","AU-9","AU-10","AU-11","AU-12"]},
    "CA": {"name": "Assessment, Authorization and Monitoring", "low": ["CA-1","CA-2","CA-3","CA-5","CA-6","CA-7","CA-9"], "moderate": ["CA-1","CA-2","CA-3","CA-5","CA-6","CA-7","CA-9"], "high": ["CA-1","CA-2","CA-3","CA-5","CA-6","CA-7","CA-8","CA-9"]},
    "CM": {"name": "Configuration Management", "low": ["CM-1","CM-2","CM-4","CM-6","CM-7","CM-8","CM-10","CM-11"], "moderate": ["CM-1","CM-2","CM-3","CM-4","CM-5","CM-6","CM-7","CM-8","CM-9","CM-10","CM-11","CM-12"], "high": ["CM-1","CM-2","CM-3","CM-4","CM-5","CM-6","CM-7","CM-8","CM-9","CM-10","CM-11","CM-12"]},
    "CP": {"name": "Contingency Planning", "low": ["CP-1","CP-2","CP-3","CP-4","CP-9","CP-10"], "moderate": ["CP-1","CP-2","CP-3","CP-4","CP-6","CP-7","CP-8","CP-9","CP-10"], "high": ["CP-1","CP-2","CP-3","CP-4","CP-6","CP-7","CP-8","CP-9","CP-10"]},
    "IA": {"name": "Identification and Authentication", "low": ["IA-1","IA-2","IA-4","IA-5","IA-6","IA-7","IA-8","IA-11"], "moderate": ["IA-1","IA-2","IA-3","IA-4","IA-5","IA-6","IA-7","IA-8","IA-11","IA-12"], "high": ["IA-1","IA-2","IA-3","IA-4","IA-5","IA-6","IA-7","IA-8","IA-11","IA-12"]},
    "IR": {"name": "Incident Response", "low": ["IR-1","IR-2","IR-4","IR-5","IR-6","IR-7","IR-8"], "moderate": ["IR-1","IR-2","IR-3","IR-4","IR-5","IR-6","IR-7","IR-8"], "high": ["IR-1","IR-2","IR-3","IR-4","IR-5","IR-6","IR-7","IR-8"]},
    "MA": {"name": "Maintenance", "low": ["MA-1","MA-2","MA-4","MA-5"], "moderate": ["MA-1","MA-2","MA-3","MA-4","MA-5"], "high": ["MA-1","MA-2","MA-3","MA-4","MA-5"]},
    "MP": {"name": "Media Protection", "low": ["MP-1","MP-2","MP-6","MP-7"], "moderate": ["MP-1","MP-2","MP-3","MP-4","MP-5","MP-6","MP-7"], "high": ["MP-1","MP-2","MP-3","MP-4","MP-5","MP-6","MP-7"]},
    "PE": {"name": "Physical and Environmental Protection", "low": ["PE-1","PE-2","PE-3","PE-6","PE-8","PE-12","PE-13","PE-14","PE-15","PE-16"], "moderate": ["PE-1","PE-2","PE-3","PE-4","PE-5","PE-6","PE-8","PE-9","PE-10","PE-11","PE-12","PE-13","PE-14","PE-15","PE-16","PE-17"], "high": ["PE-1","PE-2","PE-3","PE-4","PE-5","PE-6","PE-8","PE-9","PE-10","PE-11","PE-12","PE-13","PE-14","PE-15","PE-16","PE-17","PE-18"]},
    "PL": {"name": "Planning", "low": ["PL-1","PL-2","PL-4","PL-10","PL-11"], "moderate": ["PL-1","PL-2","PL-4","PL-10","PL-11"], "high": ["PL-1","PL-2","PL-4","PL-10","PL-11"]},
    "PM": {"name": "Program Management", "low": [], "moderate": [], "high": []},
    "PS": {"name": "Personnel Security", "low": ["PS-1","PS-2","PS-3","PS-4","PS-5","PS-6","PS-7","PS-8","PS-9"], "moderate": ["PS-1","PS-2","PS-3","PS-4","PS-5","PS-6","PS-7","PS-8","PS-9"], "high": ["PS-1","PS-2","PS-3","PS-4","PS-5","PS-6","PS-7","PS-8","PS-9"]},
    "PT": {"name": "PII Processing and Transparency", "low": [], "moderate": ["PT-1","PT-2","PT-3","PT-4"], "high": ["PT-1","PT-2","PT-3","PT-4"]},
    "RA": {"name": "Risk Assessment", "low": ["RA-1","RA-2","RA-3","RA-5","RA-7"], "moderate": ["RA-1","RA-2","RA-3","RA-5","RA-7"], "high": ["RA-1","RA-2","RA-3","RA-5","RA-7","RA-9"]},
    "SA": {"name": "System and Services Acquisition", "low": ["SA-1","SA-2","SA-3","SA-4","SA-5","SA-8","SA-9","SA-22"], "moderate": ["SA-1","SA-2","SA-3","SA-4","SA-5","SA-8","SA-9","SA-10","SA-11","SA-22"], "high": ["SA-1","SA-2","SA-3","SA-4","SA-5","SA-8","SA-9","SA-10","SA-11","SA-12","SA-15","SA-17","SA-22"]},
    "SC": {"name": "System and Communications Protection", "low": ["SC-1","SC-5","SC-7","SC-12","SC-13","SC-15","SC-20","SC-21","SC-22","SC-39"], "moderate": ["SC-1","SC-2","SC-4","SC-5","SC-7","SC-8","SC-10","SC-12","SC-13","SC-15","SC-17","SC-18","SC-20","SC-21","SC-22","SC-23","SC-28","SC-39"], "high": ["SC-1","SC-2","SC-3","SC-4","SC-5","SC-7","SC-8","SC-10","SC-12","SC-13","SC-15","SC-17","SC-18","SC-20","SC-21","SC-22","SC-23","SC-24","SC-28","SC-39"]},
    "SI": {"name": "System and Information Integrity", "low": ["SI-1","SI-2","SI-3","SI-4","SI-5","SI-12"], "moderate": ["SI-1","SI-2","SI-3","SI-4","SI-5","SI-7","SI-10","SI-12","SI-16"], "high": ["SI-1","SI-2","SI-3","SI-4","SI-5","SI-6","SI-7","SI-8","SI-10","SI-12","SI-16"]},
    "SR": {"name": "Supply Chain Risk Management", "low": [], "moderate": ["SR-1","SR-2","SR-3","SR-5","SR-6","SR-8","SR-11"], "high": ["SR-1","SR-2","SR-3","SR-5","SR-6","SR-8","SR-11"]},
}

# 800-171 families
CONTROL_FAMILIES_800_171 = {
    "3.1": {"name": "Access Control", "controls": [f"3.1.{i}" for i in range(1, 23)]},
    "3.2": {"name": "Awareness and Training", "controls": [f"3.2.{i}" for i in range(1, 4)]},
    "3.3": {"name": "Audit and Accountability", "controls": [f"3.3.{i}" for i in range(1, 10)]},
    "3.4": {"name": "Configuration Management", "controls": [f"3.4.{i}" for i in range(1, 10)]},
    "3.5": {"name": "Identification and Authentication", "controls": [f"3.5.{i}" for i in range(1, 12)]},
    "3.6": {"name": "Incident Response", "controls": [f"3.6.{i}" for i in range(1, 4)]},
    "3.7": {"name": "Maintenance", "controls": [f"3.7.{i}" for i in range(1, 7)]},
    "3.8": {"name": "Media Protection", "controls": [f"3.8.{i}" for i in range(1, 10)]},
    "3.9": {"name": "Personnel Security", "controls": [f"3.9.{i}" for i in range(1, 3)]},
    "3.10": {"name": "Physical Protection", "controls": [f"3.10.{i}" for i in range(1, 7)]},
    "3.11": {"name": "Risk Assessment", "controls": [f"3.11.{i}" for i in range(1, 4)]},
    "3.12": {"name": "Security Assessment", "controls": [f"3.12.{i}" for i in range(1, 5)]},
    "3.13": {"name": "System and Communications Protection", "controls": [f"3.13.{i}" for i in range(1, 17)]},
    "3.14": {"name": "System and Information Integrity", "controls": [f"3.14.{i}" for i in range(1, 8)]},
}


def generate_worksheet(framework, baseline, output_path, entity_name="Organization"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Control Assessment"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    family_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    family_font = Font(name="Calibri", bold=True, size=11)
    title_font = Font(name="Calibri", bold=True, size=14)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    framework_label = {
        "800-53": f"NIST SP 800-53 Rev 5 ({baseline.capitalize()} Baseline)",
        "800-171": "NIST SP 800-171 Rev 2",
    }.get(framework, framework)

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Control Assessment Worksheet — {framework_label}"
    ws["A1"].font = title_font

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Entity: {entity_name} | Date: {datetime.now().strftime('%Y-%m-%d')} | Assessor: _______________"
    ws["A2"].font = Font(size=11, italic=True)

    # Headers
    headers = [
        "Control ID", "Control Family", "Implementation\nStatus", "Effectiveness\n(0-5)",
        "Maturity\n(1-5)", "Evidence\nExamined", "Interviews\nConducted", "Tests\nPerformed",
        "Gaps / Findings", "Likelihood\n(1-5)", "Impact\n(1-5)", "Risk\nScore", "Auditor Notes",
    ]

    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    widths = [12, 22, 18, 14, 12, 20, 20, 20, 30, 12, 10, 10, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[header_row].height = 40

    # Populate controls
    row = header_row + 1

    if framework == "800-53":
        families = CONTROL_FAMILIES_800_53
        for family_id, family_data in families.items():
            controls = family_data.get(baseline, [])
            if not controls:
                continue

            # Family header row
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            cell = ws.cell(row=row, column=1, value=f"{family_id} — {family_data['name']}")
            cell.font = family_font
            cell.fill = family_fill
            row += 1

            for ctrl in controls:
                ws.cell(row=row, column=1, value=ctrl).border = thin_border
                ws.cell(row=row, column=2, value=family_data["name"]).border = thin_border
                for col in range(3, 14):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

    elif framework == "800-171":
        families = CONTROL_FAMILIES_800_171
        for family_id, family_data in families.items():
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
            cell = ws.cell(row=row, column=1, value=f"{family_id} — {family_data['name']}")
            cell.font = family_font
            cell.fill = family_fill
            row += 1

            for ctrl in family_data["controls"]:
                ws.cell(row=row, column=1, value=ctrl).border = thin_border
                ws.cell(row=row, column=2, value=family_data["name"]).border = thin_border
                for col in range(3, 14):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
                row += 1

    # Freeze panes
    ws.freeze_panes = "A5"

    # Add legend sheet
    ws_legend = wb.create_sheet("Legend")
    ws_legend["A1"] = "Assessment Legend"
    ws_legend["A1"].font = title_font

    ws_legend["A3"] = "Implementation Status Options"
    ws_legend["A3"].font = Font(bold=True)
    for i, status in enumerate(STATUS_OPTIONS):
        ws_legend.cell(row=4 + i, column=1, value=status)

    ws_legend["A12"] = "Effectiveness Rating"
    ws_legend["A12"].font = Font(bold=True)
    for i, eff in enumerate(EFFECTIVENESS_OPTIONS):
        ws_legend.cell(row=13 + i, column=1, value=eff)

    ws_legend["A21"] = "Maturity Level"
    ws_legend["A21"].font = Font(bold=True)
    for i, mat in enumerate(MATURITY_OPTIONS):
        ws_legend.cell(row=22 + i, column=1, value=mat)

    ws_legend.column_dimensions["A"].width = 40

    wb.save(output_path)
    print(f"Assessment worksheet saved to: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate NIST Control Assessment Worksheet")
    parser.add_argument("--framework", required=True, choices=["800-53", "800-171"], help="NIST framework")
    parser.add_argument("--baseline", default="moderate", choices=["low", "moderate", "high"], help="800-53 baseline (ignored for 800-171)")
    parser.add_argument("--output", default="control_assessment_worksheet.xlsx", help="Output Excel file path")
    parser.add_argument("--entity", default="Organization", help="Entity name")
    args = parser.parse_args()

    generate_worksheet(args.framework, args.baseline, args.output, args.entity)


if __name__ == "__main__":
    main()
