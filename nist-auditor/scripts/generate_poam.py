"""
Generate a Plan of Action & Milestones (POA&M) spreadsheet from audit findings.

Usage:
    python generate_poam.py --findings findings.json --output poam.xlsx

Input: JSON file with findings array
Output: Formatted Excel POA&M workbook
"""

import json
import sys
import argparse
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl required. Install with: pip install openpyxl --break-system-packages")
    sys.exit(1)


# Risk level colors
RISK_COLORS = {
    "Critical": "FF0000",
    "High": "FF6600",
    "Medium": "FFD700",
    "Low": "00B050",
}

# Remediation timeline defaults (days)
TIMELINE_DEFAULTS = {
    "Critical": 30,
    "High": 90,
    "Medium": 180,
    "Low": 365,
}


def get_risk_level(score):
    if score >= 16:
        return "Critical"
    elif score >= 10:
        return "High"
    elif score >= 5:
        return "Medium"
    else:
        return "Low"


def generate_poam(findings, output_path, entity_name="Organization", framework="NIST SP 800-53 Rev 5"):
    wb = Workbook()
    ws = wb.active
    ws.title = "POA&M"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14)
    subtitle_font = Font(name="Calibri", size=11, italic=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title rows
    ws.merge_cells("A1:K1")
    ws["A1"] = f"Plan of Action & Milestones (POA&M)"
    ws["A1"].font = title_font

    ws.merge_cells("A2:K2")
    ws["A2"] = f"Entity: {entity_name} | Framework: {framework} | Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws["A2"].font = subtitle_font

    # Headers
    headers = [
        "POA&M ID",
        "Finding",
        "Control ID(s)",
        "Control Family",
        "Weakness Description",
        "Likelihood",
        "Impact",
        "Risk Score",
        "Risk Level",
        "Status",
        "Planned Remediation",
        "Responsible Party",
        "Scheduled Completion",
        "Milestones",
        "Resources Required",
        "Comments",
    ]

    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Column widths
    widths = [10, 25, 15, 20, 40, 12, 10, 12, 12, 14, 35, 20, 18, 30, 20, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sort findings by risk score descending
    sorted_findings = sorted(findings, key=lambda f: f.get("risk_score", 0), reverse=True)

    # Data rows
    for idx, finding in enumerate(sorted_findings, 1):
        row = header_row + idx
        risk_score = finding.get("risk_score", 0)
        risk_level = finding.get("risk_level", get_risk_level(risk_score))
        likelihood = finding.get("likelihood", 0)
        impact = finding.get("impact", 0)
        default_days = TIMELINE_DEFAULTS.get(risk_level, 180)
        target_date = (datetime.now() + timedelta(days=default_days)).strftime("%Y-%m-%d")

        values = [
            f"POAM-{idx:03d}",
            finding.get("title", ""),
            finding.get("control_ids", ""),
            finding.get("control_family", ""),
            finding.get("description", ""),
            likelihood,
            impact,
            risk_score,
            risk_level,
            finding.get("status", "Open"),
            finding.get("remediation", ""),
            finding.get("responsible", "TBD"),
            finding.get("target_date", target_date),
            finding.get("milestones", ""),
            finding.get("resources", ""),
            finding.get("comments", ""),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Color-code risk level cell
        risk_cell = ws.cell(row=row, column=9)
        color = RISK_COLORS.get(risk_level, "FFFFFF")
        risk_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        if risk_level in ("Critical", "High"):
            risk_cell.font = Font(bold=True, color="FFFFFF")

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "POA&M Summary"
    ws_summary["A1"].font = title_font

    ws_summary["A3"] = "Risk Level"
    ws_summary["B3"] = "Count"
    ws_summary["C3"] = "Target Timeline"
    for cell in [ws_summary["A3"], ws_summary["B3"], ws_summary["C3"]]:
        cell.font = header_font
        cell.fill = header_fill

    levels = ["Critical", "High", "Medium", "Low"]
    for i, level in enumerate(levels):
        row = 4 + i
        count = sum(1 for f in sorted_findings if get_risk_level(f.get("risk_score", 0)) == level)
        ws_summary.cell(row=row, column=1, value=level)
        ws_summary.cell(row=row, column=2, value=count)
        ws_summary.cell(row=row, column=3, value=f"Within {TIMELINE_DEFAULTS[level]} days")
        ws_summary.cell(row=row, column=1).fill = PatternFill(
            start_color=RISK_COLORS[level], end_color=RISK_COLORS[level], fill_type="solid"
        )

    ws_summary.column_dimensions["A"].width = 15
    ws_summary.column_dimensions["B"].width = 10
    ws_summary.column_dimensions["C"].width = 20

    ws_summary.cell(row=9, column=1, value="Total Findings")
    ws_summary.cell(row=9, column=1).font = Font(bold=True)
    ws_summary.cell(row=9, column=2, value=len(sorted_findings))

    # Freeze panes
    ws.freeze_panes = "A5"

    wb.save(output_path)
    print(f"POA&M saved to: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Generate POA&M from audit findings")
    parser.add_argument("--findings", required=True, help="Path to findings JSON file")
    parser.add_argument("--output", default="poam.xlsx", help="Output Excel file path")
    parser.add_argument("--entity", default="Organization", help="Entity name")
    parser.add_argument("--framework", default="NIST SP 800-53 Rev 5", help="Framework name")
    args = parser.parse_args()

    with open(args.findings, "r") as f:
        data = json.load(f)

    findings = data if isinstance(data, list) else data.get("findings", [])
    generate_poam(findings, args.output, args.entity, args.framework)


if __name__ == "__main__":
    main()
